import re
import openpyxl

from fastapi import APIRouter, UploadFile
from src.config import path_main
from src.payments.routers.re_pattern_pay import RePattern


'''
Метод извлечения расходов из банковских выписок Excel
'''
router_extract_expenses = APIRouter(
    prefix="/v1/ExtractExpensesExcel",
    tags=["Payments"]
)


@router_extract_expenses.post("/")
async def extract_expenses(file_object: UploadFile):

    with open(f'{path_main}/src/media/data/bank_records_file.xlsx', 'wb+') as f:
        for chunk in file_object.file:
            f.write(chunk)

    path_file = f'{path_main}/src/media/data/bank_records_file.xlsx'

    extract_expenses_list = expenses_from_bank(path_file)

    data_payment_1 = []
    summa_pay = 0
    count_pay = 0

    sd_date_pay = None
    sd_summa_pay = 0
    sd_purpose_pay = None

    fuel_date_pay = None
    fuel_summa_pay = 0
    fuel_purpose_pay = None

    bank_date_pay = None
    bank_summa_pay = 0
    bank_purpose_pay = None


    for pay in extract_expenses_list:
        if 'Госпошлина' in pay['purpose_pay']:
            sd_date_pay = pay['date']
            sd_summa_pay += pay['summ_pay']
            sd_purpose_pay = pay['purpose_pay']
            count_pay += 1
            summa_pay += pay['summ_pay']
        elif 'Лукойл' in pay['purpose_pay']:
            fuel_date_pay = pay['date']
            fuel_summa_pay += pay['summ_pay']
            fuel_purpose_pay = pay['purpose_pay']
            count_pay += 1
            summa_pay += pay['summ_pay']
        elif 'Альфа-банк' in pay['purpose_pay']:
            bank_date_pay = pay['date']
            bank_summa_pay += pay['summ_pay']
            bank_purpose_pay = pay['purpose_pay']
            count_pay += 1
            summa_pay += pay['summ_pay']
        else:
            count_pay += 1
            summa_pay += pay['summ_pay']

            data_payment_1.append({
                "id": None,
                "cession_id": None,
                "cession": 'Не определен',
                "expenses_category_id": None,
                "expenses_category": None,
                "datePay": pay['date'],
                "summaPay": pay['summ_pay'],
                "purposePay": pay['purpose_pay'],
            })

    data_payment_2 = [
        {
            "id": None,
            "cession_id": None,
            "cession": 'Не определен',
            "expenses_category_id": None,
            "expenses_category": None,
            "datePay": sd_date_pay,
            "summaPay": round(sd_summa_pay, 2),
            "purposePay": sd_purpose_pay,
        },
        {
            "id": None,
            "cession_id": None,
            "cession": 'Не определен',
            "expenses_category_id": None,
            "expenses_category": None,
            "datePay": fuel_date_pay,
            "summaPay": round(fuel_summa_pay, 2),
            "purposePay": fuel_purpose_pay,
        },
        {
            "id": None,
            "cession_id": None,
            "cession": 'Не определен',
            "expenses_category_id": None,
            "expenses_category": None,
            "datePay": bank_date_pay,
            "summaPay": round(bank_summa_pay, 2),
            "purposePay": bank_purpose_pay,
        }]

    result = {'data_payment': data_payment_1 + data_payment_2,
              'summa_all': round(summa_pay, 2),
              'count_all': count_pay}

    return result


def expenses_from_bank(path_file):

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    payments = []

    col_date_payment = 0
    col_debt = 0
    col_num_doc = 0
    col_purpose = 0
    for row in range(1, sheet.max_row + 1):
        for cell in range(0, sheet.max_column):
            if re.findall(r'^Дата$', str(sheet[row][cell].value)):
                col_date_payment = cell
                # print(f'{col_date_payment=}')
            if re.findall(r'^Номер\s+документа$', str(sheet[row][cell].value)):
                col_num_doc = cell
                # print(f'{col_num_doc=}')
            if re.findall(r'^Дебет$', str(sheet[row][cell].value)):
                col_debt = cell
                # print(f'{col_debt=}')
            if re.findall(r'^Назначение\s+платежа$', str(sheet[row + 1][cell].value)):
                col_purpose = cell


        # Если столбец "Дебет" содержит данные, то парсинг строки
        if re.findall(RePattern.payment, str(sheet[row][col_debt].value)):
            payment = sheet[row][col_debt].value
            purpose_payment = sheet[row][col_purpose].value
            num_payment = sheet[row][col_num_doc].value
            if re.findall(r'\d{2}\.\d{2}\.\d{4}', str(sheet[row][col_date_payment].value)):
                date_payment = re.sub(r'\n', '', sheet[row][col_date_payment].value)
                # print(date_payment)
            else:
                date_format = sheet[row][col_date_payment].value
                date_payment = date_format.strftime("%d.%m.%Y")

            if re.findall(RePattern.state_duty, purpose_payment):
                purpose_pay = 'Госпошлина'
            elif re.findall(RePattern.card_pay, purpose_payment):
                card_pay = purpose_payment
                if re.findall(RePattern.lukoil_pay, card_pay):
                    purpose_pay = 'Транспортные расходы - Лукойл'
                else:
                    purpose_pay = 'Оплата по карте Альфа-банк'
            else:
                purpose_pay = purpose_payment

            payments.append({"date": date_payment,
                             "summ_pay": payment,
                             "purpose_pay": purpose_pay})

    return payments