import re
import openpyxl

from fastapi import APIRouter, UploadFile, Form
from src.config import path_main
from src.payments.routers.re_pattern_pay import RePattern


'''
Метод извлечения платежей из банковских выписок Excel
'''
router_extract_payments = APIRouter(
    prefix="/v1/ExtractPaymentsExcel",
    tags=["Payments"]
)


@router_extract_payments.post("/")
async def extract_payments(file_object: UploadFile, checkFunction: str = Form(...)):

    with open(f'{path_main}/src/media/data/bank_records_file.xlsx', 'wb+') as f:
        for chunk in file_object.file:
            f.write(chunk)

    path_file = f'{path_main}/src/media/data/bank_records_file.xlsx'

    if checkFunction == '1':
        extract_payments_list = payment_from_bank(path_file)
    else:
        extract_payments_list = refund_payments_from_cedent(path_file)

    data_payment = []
    summa_pay = 0
    count_pay = 0

    for pay in extract_payments_list:
        credit_id = None
        credit_num = 'Отсутствует'
        count_pay += 1
        summa_pay += pay['payment']

        data_payment.append({
            "credit_id": credit_id,
            "creditNum": credit_num,
            "debtorName": pay['fio'],
            "numEP": pay['ep'],
            "numED": pay['ed'],
            "summa": pay['payment'],
            "date": pay['date'],
            "numPayDoc": pay['num_pay'],
            "purposePay": pay['purpose_pay'],
        })

    result = {'data_payment': data_payment,
              'summa_all': summa_pay,
              'count_all': count_pay}

    return result


def payment_from_bank(path_file):

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    payments = []

    col_date_payment = 0
    col_credit = 0
    col_num_doc = 0
    col_purpose = 0
    for row in range(1, sheet.max_row + 1):
        for cell in range(0, sheet.max_column):
            if re.findall(r'^Дата$', str(sheet[row][cell].value)):
                col_date_payment = cell
                # print(f'{col_date_payment=}')
            if re.findall(r'^Номер\s+документа$', str(sheet[row][cell].value)):
                col_num_doc = cell
                # print(f'{col_num_doc=}')
            if re.findall(r'^Кредит$', str(sheet[row][cell].value)):
                col_credit = cell
                # print(f'{col_credit=}')
            if re.findall(r'^Назначение\s+платежа$', str(sheet[row + 1][cell].value)):
                col_purpose = cell


        # Если столбец "Кредит" содержит данные, то парсинг строки
        if re.findall(RePattern.payment, str(sheet[row][col_credit].value)):
            payment_in = sheet[row][col_credit].value
            purpose_payment = sheet[row][col_purpose].value
            num_payment = sheet[row][col_num_doc].value
            # print(num_payment)
            if re.findall(r'\d{2}\.\d{2}\.\d{4}', str(sheet[row][col_date_payment].value)):
                date_payment = re.sub(r'\n', '', sheet[row][col_date_payment].value)
                # print(date_payment)
            else:
                date_format = sheet[row][col_date_payment].value
                # print(date_format)
                date_payment = date_format.strftime("%d.%m.%Y")

            if re.findall(r'(?i)(УФК)', purpose_payment):
                try:
                    fio_debtor = re.search(RePattern.pay_fio_ufk, purpose_payment).group()
                except:
                    fio_debtor = 'ФИО-ОШИБКА_УФК'
            elif re.findall(r'(?i)(ОСФР)', purpose_payment):
                try:
                    fio_debtor = re.search(RePattern.pay_fio_osfr, purpose_payment).group()
                except:
                    fio_debtor = 'ФИО-ОШИБКА_ОСФР'
            else:
                try:
                    fio_debtor = re.search(RePattern.pay_fio, purpose_payment).group()
                except:
                    fio_debtor = 'ФИО-ОШИБКА'

            purpose_pay = purpose_payment

            # if re.findall(r'(?i)\(', purpose_payment):
            #     try:
            #         department = re.search(RePattern.department_ufk, purpose_payment).group()
            #     except:
            #         department = 'Департамент_ОШИБКА'
            # elif re.findall(r'(?i)(Банк)', purpose_payment):
            #     try:
            #         department = re.search(RePattern.department_bank, purpose_payment).group()
            #     except:
            #         department = purpose_payment
            # else:
            #     try:
            #         department = re.search(RePattern.pay_fio_bank, purpose_payment).group()
            #     except:
            #         department = purpose_payment

            try:
                executive_production = re.search(RePattern.number_case, purpose_payment).group()
            except:
                executive_production = ''

            try:
                executive_document = re.search(RePattern.executive_document, purpose_payment).group()
            except:
                executive_document = ''

            try:
                ispol_list = re.search(RePattern.ispol_list, purpose_payment).group()
            except:
                ispol_list = ''

            payments.append({"fio": fio_debtor,
                             "ep": executive_production,
                             "ed": executive_document,
                             "il": ispol_list,
                             "date": date_payment,
                             "payment": payment_in,
                             "num_pay": num_payment,
                             "purpose_pay": purpose_pay})

    return payments


def refund_payments_from_cedent(path_file):
    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    payments = []

    col_result = 0
    col_debit = 0
    col_summa = 0
    date_pay = None
    purpose_pay = 'Возврат платежей от Цедента'
    for row in range(1, sheet.max_row):
        for cell in range(0, sheet.max_column):

            if re.findall(r'(?i)итого', str(sheet[row][cell].value)):
                col_result = cell
            if re.findall(r'по\s+дебету', str(sheet[row][cell].value)):
                col_debit = cell
            if re.findall(r'в\s+рублях', str(sheet[row][cell].value)):
                col_summa = cell
            if re.findall(r'(?i)Дата\s+операционного\s+дня', str(sheet[row][cell].value)):
                date_pay = re.search(r'\d{2}\.\d{2}\.\d{4}', str(sheet[row][cell].value)).group()

        # Если столбец "Поступление" содержит данные, то парсинг строки
        if re.findall(r'(?i)итого', str(sheet[row][col_result].value)):
            return payments

        elif re.findall(RePattern.payment_memori, str(sheet[row][col_summa].value)):
            payment_text = sheet[row][col_summa].value
            try:
                payment_format = re.sub(r'-', '.', payment_text)
            except:
                payment_format = payment_text
            payment = float(payment_format)

            debtor = re.search(RePattern.pay_fio_memori, str(sheet[row][col_debit].value)).group()

            payments.append({"fio": debtor,
                             "ep": None,
                             "ed": None,
                             "il": None,
                             "date": date_pay,
                             "payment": payment,
                             "num_pay": None,
                             "purpose_pay": purpose_pay})

    return payments