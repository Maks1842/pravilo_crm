import re
import openpyxl
from fastapi import APIRouter, UploadFile, Depends
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from src.database import get_async_session
from src.config import path_main

from src.mail.models import mail_out


'''
Метод извлечения платежей из банковских выписок Excel
'''
router_extract_mail_barcode = APIRouter(
    prefix="/v1/ExtractMailBarcodeExcel",
    tags=["Reader"]
)


@router_extract_mail_barcode.post("/")
async def extract_mail_barcode(file_object: UploadFile, session: AsyncSession = Depends(get_async_session)):

    with open(f'{path_main}/src/media/data/mail_barcode_file.xlsx', 'wb+') as f:
        for chunk in file_object.file:
            f.write(chunk)

    path_file = f'{path_main}/src/media/data/mail_barcode_file.xlsx'

    extract_barcode = mail_barcode_reader_xlsx(path_file)

    data_mail = []
    count_mail = 0

    for item in extract_barcode:

        mail_out_query = await session.execute(select(mail_out.c.id).where(mail_out.c.sequence_num == int(item['mail_number'])))
        mail_out_id = mail_out_query.scalar()

        if mail_out_id:
            mail_number = item['mail_number']
        else:
            mail_number = 'Не определен'

        count_mail += 1

        data_mail.append({
            "mail_out_id": mail_out_id,
            "mail_out_number": mail_number,
            "mail_order_number": item['mail_order_number'],
            "order_barcode": item['order_barcode'],
        })

    result = {'data_mail': data_mail,
              'count_all': count_mail}

    return result


def mail_barcode_reader_xlsx(path_file):

    mails = []

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    col_order_number = 0
    col_order_barcode = 0

    for cell in range(1, sheet.max_column):
        if re.findall(r'(?i)Номер\s+заказа', str(sheet[1][cell].value)):
            col_order_number = cell
        if re.findall(r'(?i)ШПИ', str(sheet[1][cell].value)):
            col_order_barcode = cell

    for row in range(2, sheet.max_row + 1):
        order_number = str(sheet[row][col_order_number].value)
        order_barcode = str(sheet[row][col_order_barcode].value)

        try:
            mail_number = re.sub(r'\s+', '', re.search(r'(?<=№)\s+\d+\s+(?=от)', order_number).group())

        except:
            mail_number = 'ОШИБКА'

        mails.append({"mail_number": mail_number,
                      "mail_order_number": order_number,
                      "order_barcode": order_barcode})

    return mails