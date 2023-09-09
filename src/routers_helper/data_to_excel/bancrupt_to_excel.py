from openpyxl import load_workbook
from datetime import datetime

from src.routers_helper.data_to_excel.style_excel import style_excel

def bankrupt_to_excel(data_list):

    style = style_excel()

    book_template = load_workbook(filename='/media/maks/Новый том/Python/work/fast_api/pravilo_crm/src/media/bankrupt/template_bankrupt.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Банкроты"

    number_col = 1
    number_row = 2
    count = 0
    for data in data_list:
        count += 1
        number_col += 1

        # # Вставить один столбец перед №___
        # sheet.insert_cols(number_col)

        column = 1
        row = number_row

        # Высота строки
        sheet.row_dimensions[row].height = 25

        # Ширина столбца
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 60
        sheet.column_dimensions["C"].width = 60
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 70
        sheet.column_dimensions["F"].width = 70
        sheet.column_dimensions["G"].width = 20
        sheet.column_dimensions["H"].width = 20
        sheet.column_dimensions["I"].width = 30
        sheet.column_dimensions["J"].width = 70

        try:
            # Добавить данные в ячейку
            column = 1
            sheet.cell(row, column).value = count
            sheet.cell(row, column).style = style['style_main']

            column = 2
            sheet.cell(row, column).value = data['full_name']
            sheet.cell(row, column).style = style['style_main']

            column = 3
            sheet.cell(row, column).value = data['name_histories']
            sheet.cell(row, column).style = style['style_main']

            column = 4
            date_conv = datetime.strptime(data['birthdate_bankrupt'], '%Y-%m-%dT00:00:00').strftime('%d.%m.%Y')
            sheet.cell(row, column).value = date_conv
            sheet.cell(row, column).style = style['style_main']

            column = 5
            sheet.cell(row, column).value = data['birth_place_bankrupt']
            sheet.cell(row, column).style = style['style_main']

            column = 6
            sheet.cell(row, column).value = data['address']
            sheet.cell(row, column).style = style['style_main']

            column = 7
            sheet.cell(row, column).value = data['inn']
            sheet.cell(row, column).style = style['style_main']

            column = 8
            sheet.cell(row, column).value = data['snils']
            sheet.cell(row, column).style = style['style_main']

            column = 9
            sheet.cell(row, column).value = data['number_legal_cases']
            sheet.cell(row, column).style = style['style_main']

            column = 10
            sheet.cell(row, column).value = data['tribunal']
            sheet.cell(row, column).style = style['style_main']
        except Exception as ex:
            return {
                "status": "error",
                "data": None,
                "details": f'Реестр банкротов не сформирован, ошибка на строке {row}. {ex}'
            }

        number_row += 1

    current_date = datetime.now()
    file = f'/media/maks/Новый том/Python/work/fast_api/pravilo_crm/src/media/bankrupt/result/Реестр банкротов_{current_date.strftime("%d.%m.%Y")}.xlsx'
    book_template.save(file)

    return {
        'status': 'success',
        'data': None,
        'details': 'Реестр банкротов успешно сформирован'
    }