from openpyxl import load_workbook
from datetime import datetime, timedelta

from src.routers_helper.data_to_excel.style_excel import style_excel
from src.config import path_main

def calculating_debt_to_excel(data):

    date_start = data['date_start_cd']
    date_start_format = datetime.strptime(str(date_start), '%Y-%m-%d').strftime("%d.%m.%Y")

    cession_date = datetime.strptime(str(data['cession_date']), '%Y-%m-%d').strftime("%d.%m.%Y")

    current_date = datetime.now()
    current_date_format = current_date.strftime("%d.%m.%Y")

    summa_stop = data['summa_cd'] * 1.5

    style = style_excel()

    book_template = load_workbook(filename=f'{path_main}/src/media/calculating_debt/template_calculating_debt.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "График"

    # # Вставить один столбец перед №___
    # sheet.insert_cols(number_col)

    # Высота строки
    # sheet.row_dimensions[row].height = 50

    # Ширина столбца
    # sheet.column_dimensions["A"].width = 60

    try:
        # Добавить данные в ячейку
        row = 4
        column = 3
        sheet.cell(row, column).value = data['fio']
        # sheet.cell(row, column).style = style['style_main']

        row = 5
        column = 5
        sheet.cell(row, column).value = f"{data['number_cd']} от {date_start_format}"
        # sheet.cell(row, column).style = style['style_main']
        # print('2')

        row = 6
        column = 3
        sheet.cell(row, column).value = f"{data['summa_cd']} рублей"
        # sheet.cell(row, column).style = style['style_main']
        # print('3')

        row = 7
        column = 4
        sheet.cell(row, column).value = f"{data['interest_rate']}%"
        # sheet.cell(row, column).style = style['style_main']

        row = 8
        column = 4
        sheet.cell(row, column).value = f"(до {data['date_end']} включительно)"


        summa_percent_all = 0
        number_row = 13

        while summa_percent_all < summa_stop:

            row = number_row

            date_second = date_start + timedelta(days=30)
            date_second_formar = datetime.strptime(str(date_second), '%Y-%m-%d').strftime("%d.%m.%Y")

            summa_percent = round(data['summa_cd'] * 30 / 365 * int(data['interest_rate']) / 100, 2)
            summa_percent_all += summa_percent

            column = 1
            sheet.cell(row, column).value = data['summa_cd']
            sheet.cell(row, column).style = style['style_7']

            column = 2
            sheet.cell(row, column).value = date_start_format
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = date_second_formar
            sheet.cell(row, column).style = style['style_7']

            column = 4
            sheet.cell(row, column).value = 30
            sheet.cell(row, column).style = style['style_7']

            column = 5
            sheet.cell(row, column).value = f"{data['summa_cd']}*30/365*{data['interest_rate']}%"
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row, column).value = summa_percent
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row + 1, column).value = summa_percent_all
            sheet.cell(row + 1, column).style = style['style_7']

            number_row += 1
            date_start = date_second + timedelta(days=1)
            date_start_format = datetime.strptime(str(date_start), '%Y-%m-%d').strftime("%d.%m.%Y")

        row = row + 3
        column = 2
        sheet.cell(row, column).value = f"Общая сумма задолженности по состоянию на {cession_date} составляет:"
        sheet.cell(row, column).style = style['style_9']

        column = 2
        sheet.cell(row + 2, column).value = f"Сумма основного долга - {data['summa_cd']}  руб."
        sheet.cell(row + 2, column).style = style['style_8']
        column = 2
        sheet.cell(row + 3, column).value = f"Сумма процентов - {round(summa_percent_all, 2)}  руб."
        sheet.cell(row + 3, column).style = style['style_9']
        column = 2
        sheet.cell(row + 4, column).value = f"ИТОГО - {data['summa_cd'] + summa_percent_all}  руб."
        sheet.cell(row + 4, column).style = style['style_9']

        column = 2
        sheet.cell(row + 7, column).value = f"Генеральный директор"
        sheet.cell(row + 7, column).style = style['style_8']
        column = 2
        sheet.cell(row + 8, column).value = f"ООО «ИКЦ «Правило»"
        sheet.cell(row + 8, column).style = style['style_9']
        column = 6
        sheet.cell(row + 8, column).value = f"Симонова Ю.В."
        sheet.cell(row + 8, column).style = style['style_9']
    except Exception as ex:
        return {
            "status": "error",
            "data": None,
            "details": f'Почтовый реестр не сформирован. Предоставлены не все данные. Ошибка на строке {ex}'
        }

    file = f'{path_main}/src/media/calculating_debt/result/Расчет задолженности_{data["fio"]}_{current_date.strftime("%d.%m.%Y_%H.%M.%S")}.xlsx'
    book_template.save(file)

    return {
        'status': 'success',
        'data': None,
        'details': 'Почтовый реестр Excel успешно сформирован'
    }


def calculating_annuity_to_excel(data):

    print(f'{data=}')

    summ_percent_total = data["summ_percent_total"]

    date_start = data['date_start_cd']
    date_start_format = datetime.strptime(str(date_start), '%Y-%m-%d').strftime("%d.%m.%Y")

    cession_date = datetime.strptime(str(data['cession_date']), '%Y-%m-%d').strftime("%d.%m.%Y")

    current_date = datetime.now()

    timeline = int(data['timeline'])
    summa_stop = data['summa_cd'] * 1.5


    new_date_calculat_law = datetime.strptime('2023-07-01', '%Y-%m-%d').strftime("%d.%m.%Y")

    if date_start_format <= new_date_calculat_law:
        summa_stop = data['summa_cd'] * 1.3

    style = style_excel()

    book_template = load_workbook(filename=f'{path_main}/src/media/calculating_debt/template_calculating_annuity.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "График"

    # # Вставить один столбец перед №___
    # sheet.insert_cols(number_col)

    # Высота строки
    # sheet.row_dimensions[row].height = 50

    # Ширина столбца
    # sheet.column_dimensions["A"].width = 60

    try:
        # Добавить данные в ячейку
        row = 4
        column = 3
        sheet.cell(row, column).value = data['fio']

        row = 5
        column = 5
        sheet.cell(row, column).value = f"{data['number_cd']} от {date_start_format}"

        row = 6
        column = 3
        sheet.cell(row, column).value = f"{data['summa_cd']} рублей"

        row = 7
        column = 4
        sheet.cell(row, column).value = f"{data['interest_rate']} %, годовых"

        row = 8
        column = 3
        sheet.cell(row, column).value = f"(до {data['date_end']} включительно)"

        row = 13
        column = 1
        sheet.cell(row, column).value = data['summa_cd']
        sheet.cell(row, column).style = style['style_7']
        sheet.cell(row, column).number_format = '0.00'

        sheet.cell(row, 2).style = style['style_7']
        sheet.cell(row, 3).style = style['style_7']
        sheet.cell(row, 4).style = style['style_7']
        sheet.cell(row, 5).style = style['style_7']
        sheet.cell(row, 6).style = style['style_7']
        sheet.cell(row, 7).style = style['style_7']
        sheet.cell(row, 8).style = style['style_7']
        sheet.cell(row, 9).style = style['style_7']

        row = 14

        summa_percent_all = 0
        summa_perc_second = 0
        date_start_2 = None
        if data['date_end_pay']:
            annuity_list = data['result']
            for item in annuity_list:

                date_pay = item['date_pay'].date().strftime("%d.%m.%Y")
                pay_od = item['pay_od']
                pay_percent = item['pay_percent']
                remains_debt = item['remains_debt']

                column = 1
                sheet.cell(row, column).value = remains_debt
                sheet.cell(row, column).style = style['style_7']
                sheet.cell(row, column).number_format = '0.00'

                column = 2
                sheet.cell(row, column).value = pay_od
                sheet.cell(row, column).style = style['style_7']
                sheet.cell(row, column).number_format = '0.00'

                column = 3
                sheet.cell(row, column).value = pay_percent
                sheet.cell(row, column).style = style['style_7']
                sheet.cell(row, column).number_format = '0.00'

                column = 4
                sheet.cell(row, column).value = date_pay
                sheet.cell(row, column).style = style['style_7']

                sheet.cell(row, 5).style = style['style_7']
                sheet.cell(row, 6).style = style['style_7']
                sheet.cell(row, 7).style = style['style_7']
                sheet.cell(row, 8).style = style['style_7']
                sheet.cell(row, 9).style = style['style_7']

                date_start_2 = item['date_pay']

                row += 1
        else:
            date_start_pay = datetime.strptime(str(data['date_start_pay']), '%Y-%m-%d').strftime("%d.%m.%Y")

            column = 1
            sheet.cell(row, column).value = data['summa_cd']
            sheet.cell(row, column).style = style['style_7']
            sheet.cell(row, column).number_format = '0.00'

            column = 5
            sheet.cell(row, column).value = date_start_format
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row, column).value = date_start_pay
            sheet.cell(row, column).style = style['style_7']

            column = 7
            sheet.cell(row, column).value = timeline
            sheet.cell(row, column).style = style['style_7']

            column = 8
            sheet.cell(row, column).value = f"{data['overdue_od']}x{timeline}/365x{data['interest_rate']}%"
            sheet.cell(row, column).style = style['style_7']

            column = 9
            sheet.cell(row, column).value = data['summa_percent_start']
            sheet.cell(row, column).style = style['style_7']
            sheet.cell(row, column).number_format = '0.00'

            summa_perc_second = summ_percent_total = data['summa_percent_start']
            date_start_2 = datetime.strptime(data['date_start_pay'], '%Y-%m-%d')

            row += 1

        date_start_2 = date_start_2 + timedelta(days=1)
        date_start_2_formar = date_start_2.date().strftime("%d.%m.%Y")

        while summa_percent_all < summa_stop:
            sheet.cell(row, 2).style = style['style_7']
            sheet.cell(row, 3).style = style['style_7']
            sheet.cell(row, 4).style = style['style_7']

            date_second = date_start_2 + timedelta(days=timeline - 1)
            date_second_formar = date_second.date().strftime("%d.%m.%Y")

            summa_percent = round(data['overdue_od'] * timeline / 365 * int(data['interest_rate']) / 100, 2)

            summa_perc_second += summa_percent
            summa_percent_all = summa_perc_second + summ_percent_total

            column = 1
            sheet.cell(row, column).value = data['overdue_od']
            sheet.cell(row, column).style = style['style_7']
            sheet.cell(row, column).number_format = '0.00'

            column = 5
            sheet.cell(row, column).value = date_start_2_formar
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row, column).value = date_second_formar
            sheet.cell(row, column).style = style['style_7']

            column = 7
            sheet.cell(row, column).value = timeline
            sheet.cell(row, column).style = style['style_7']

            column = 8
            sheet.cell(row, column).value = f"{data['overdue_od']}x{timeline}/365x{data['interest_rate']}%"
            sheet.cell(row, column).style = style['style_7']

            column = 9
            sheet.cell(row, column).value = summa_percent
            sheet.cell(row, column).style = style['style_7']
            sheet.cell(row, column).number_format = '0.00'

            column = 9
            sheet.cell(row + 1, column).value = summa_perc_second
            sheet.cell(row + 1, column).style = style['style_7']
            sheet.cell(row + 1, column).number_format = '0.00'

            row += 1
            date_start_2 = date_second + timedelta(days=1)
            date_start_2_formar = date_start_2.date().strftime("%d.%m.%Y")

        row = row + 3
        column = 2
        sheet.cell(row, column).value = f"Общая сумма задолженности по состоянию на {cession_date} составляет:"
        sheet.cell(row, column).style = style['style_9']

        column = 2
        sheet.cell(row + 2, column).value = f"Сумма основного долга - {data['overdue_od']}  руб."
        sheet.cell(row + 2, column).style = style['style_8']
        column = 2
        sheet.cell(row + 3, column).value = f"Сумма процентов - {round(summa_perc_second, 2)}  руб."
        sheet.cell(row + 3, column).style = style['style_9']
        column = 2
        sheet.cell(row + 4, column).value = f"ИТОГО - {round(data['overdue_od'] + summa_perc_second, 2)}  руб."
        sheet.cell(row + 4, column).style = style['style_9']

        column = 2
        sheet.cell(row + 7, column).value = f"Генеральный директор"
        sheet.cell(row + 7, column).style = style['style_8']
        column = 2
        sheet.cell(row + 8, column).value = f"ООО «ИКЦ «Правило»"
        sheet.cell(row + 8, column).style = style['style_9']
        column = 6
        sheet.cell(row + 8, column).value = f"Симонова Ю.В."
        sheet.cell(row + 8, column).style = style['style_9']
    except Exception as ex:
        return {
            "status": "error",
            "data": None,
            "details": f'Ошибка расчета графика платежей. {ex}'
        }

    file = f'{path_main}/src/media/calculating_debt/result/Аннуитет/Расчет задолженности_{data["fio"]}_{current_date.strftime("%d.%m.%Y_%H.%M.%S")}.xlsx'
    book_template.save(file)

    return {
        'status': 'success',
        'data': None,
        'details': 'График платежей успешно сформирован'
    }