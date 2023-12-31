from openpyxl import load_workbook
from datetime import datetime

from src.routers_helper.data_to_excel.style_excel import style_excel
from src.config import path_main


def statistic_to_excel_organisation(data):

    style = style_excel()

    current_date = datetime.now()

    book_template = load_workbook(filename=f'{path_main}/src/media/statistic/template_statistic_org.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Статистика"

    # # Вставить один столбец перед №___
    # sheet.insert_cols(number_col)

    # Высота строки
    # sheet.row_dimensions[row].height = 50

    # Ширина столбца
    # sheet.column_dimensions["A"].width = 60

    period = "Весь период"
    if data['date_1'] != '' and data['date_1'] != None:
        period = f"с {data['date_1']} по {data['date_2']}"

    try:
        # Добавить данные в ячейку
        row = 1
        column = 5
        sheet.cell(row, column).value = period
        # sheet.cell(row, column).style = style['style_main']

        row = 4
        column = 3
        sheet.cell(row, column).value = data['payment_total']
        # sheet.cell(row, column).style = style['style_main']
        # print('2')

        row = 5
        column = 3
        sheet.cell(row, column).value = data['expenses_total']
        # sheet.cell(row, column).style = style['style_main']
        # print('3')

        row = 6
        column = 3
        sheet.cell(row, column).value = data['accrual_expenses_total']

        row = 7
        column = 3
        sheet.cell(row, column).value = data['profit_total']
        # sheet.cell(row, column).style = style['style_main']


        row = 11
        count = 1
        for item in data['data_statistic']:

            column = 1
            sheet.cell(row, column).value = count
            sheet.cell(row, column).style = style['style_7']

            column = 2
            sheet.cell(row, column).value = item['cession_name']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item['coefficient_cession']
            sheet.cell(row, column).style = style['style_7']

            column = 4
            sheet.cell(row, column).value = item['cession_number_credit_total']
            sheet.cell(row, column).style = style['style_7']

            column = 5
            sheet.cell(row, column).value = item['cession_number_credit']
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row, column).value = item['summa_pay_cess']
            sheet.cell(row, column).style = style['style_7']

            column = 7
            sheet.cell(row, column).value = item['summa_expenses_cess']
            sheet.cell(row, column).style = style['style_7']

            column = 8
            sheet.cell(row, column).value = item['summa_accrual_cess']
            sheet.cell(row, column).style = style['style_7']

            row += 1
            count += 1

        row = row + 2

        column = 2
        sheet.cell(row, column).value = f"Расшифровка расходов за отчет.период:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 1
        column = 2
        sheet.cell(row, column).value = "Категория"
        sheet.cell(row, column).style = style['style_1_1']

        column = 3
        sheet.cell(row, column).value = "Сумма"
        sheet.cell(row, column).style = style['style_1_1']

        row = row + 1
        for item_exp in data['data_expenses_category']:

            column = 2
            sheet.cell(row, column).value = item_exp['category']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item_exp['category_summa']
            sheet.cell(row, column).style = style['style_7']

            row += 1

        row = row + 1

        column = 2
        sheet.cell(row, column).value = f"Расшифровка начислений за отчет.период:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 1
        column = 2
        sheet.cell(row, column).value = "Категория"
        sheet.cell(row, column).style = style['style_1_1']

        column = 3
        sheet.cell(row, column).value = "Сумма"
        sheet.cell(row, column).style = style['style_1_1']

        row = row + 1
        for item_exp in data['data_accrual_expenses_category']:

            column = 2
            sheet.cell(row, column).value = item_exp['category']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item_exp['category_accrual_summa']
            sheet.cell(row, column).style = style['style_7']

            row += 1
    except Exception as ex:
        return {
            "status": "error",
            "data": None,
            "details": f'Ошибка формирования отчета Excel. Ошибка на строке {ex}'
        }

    file = f'{path_main}/src/media/statistic/result/Статистика_за период с {data["date_1"]} по {data["date_2"]}_{current_date.strftime("%H%M%S")}.xlsx'
    book_template.save(file)

    return {
        'status': 'success',
        'data': None,
        'details': 'Отчет Excel успешно сформирован'
    }


def statistic_to_excel_investor(data):
    style = style_excel()

    current_date = datetime.now()

    book_template = load_workbook(filename=f'{path_main}/src/media/statistic/template_statistic_investor.xlsx')

    # Первый (активный) Лист книги
    sheet = book_template.active

    # Изменить имя Листа книги
    sheet.title = "Статистика"

    # # Вставить один столбец перед №___
    # sheet.insert_cols(number_col)

    # Высота строки
    # sheet.row_dimensions[row].height = 50

    # Ширина столбца
    # sheet.column_dimensions["A"].width = 60

    period = "Весь период"
    if data['date_1'] != '' and data['date_1'] != None:
        period = f"с {data['date_1']} по {data['date_2']}"

    try:
        # Добавить данные в ячейку
        row = 1
        column = 3
        sheet.cell(row, column).value = period
        # sheet.cell(row, column).style = style['style_main']

        row = 11
        column = 3
        sheet.cell(row, column).value = data['payment_total']
        # sheet.cell(row, column).style = style['style_main']
        # print('2')

        row = 12
        column = 3
        sheet.cell(row, column).value = data['expenses_total']
        # sheet.cell(row, column).style = style['style_main']
        # print('3')

        row = 13
        column = 3
        sheet.cell(row, column).value = data['accrual_expenses_total']
        # sheet.cell(row, column).style = style['style_main']

        row = 14
        column = 3
        sheet.cell(row, column).value = data['profit_total']
        # sheet.cell(row, column).style = style['style_main']


        row = 18
        count = 1
        for item in data['data_statistic']:

            column = 1
            sheet.cell(row, column).value = count
            sheet.cell(row, column).style = style['style_7']

            column = 2
            sheet.cell(row, column).value = item['cession_name']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item['cession_number_credit_total']
            sheet.cell(row, column).style = style['style_7']

            column = 4
            sheet.cell(row, column).value = item['cession_number_credit']
            sheet.cell(row, column).style = style['style_7']

            column = 5
            sheet.cell(row, column).value = item['summa_pay_cess']
            sheet.cell(row, column).style = style['style_7']

            column = 6
            sheet.cell(row, column).value = item['summa_expenses_cess']
            sheet.cell(row, column).style = style['style_7']

            column = 7
            sheet.cell(row, column).value = item['summa_accrual_cess']
            sheet.cell(row, column).style = style['style_7']

            row += 1
            count += 1

        row = row + 2
        column = 2
        sheet.cell(row, column).value = f"Расшифровка расходов за отчет.период:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 1
        column = 2
        sheet.cell(row, column).value = "Категория"
        sheet.cell(row, column).style = style['style_1_1']

        column = 3
        sheet.cell(row, column).value = "Сумма"
        sheet.cell(row, column).style = style['style_1_1']

        row = row + 1
        for item_exp in data['data_expenses_category']:

            column = 2
            sheet.cell(row, column).value = item_exp['category']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item_exp['category_summa']
            sheet.cell(row, column).style = style['style_7']

            row += 1

        row = row + 1
        column = 2
        sheet.cell(row, column).value = f"Расшифровка начислений за отчет.период:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 1
        column = 2
        sheet.cell(row, column).value = "Категория"
        sheet.cell(row, column).style = style['style_1_1']

        column = 3
        sheet.cell(row, column).value = "Сумма"
        sheet.cell(row, column).style = style['style_1_1']

        row = row + 1
        for item_exp in data['data_accrual_expenses_category']:

            column = 2
            sheet.cell(row, column).value = item_exp['category']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item_exp['category_accrual_summa']
            sheet.cell(row, column).style = style['style_7']

            row += 1

        row = row + 1
        column = 2
        sheet.cell(row, column).value = f"Дивиденды инвестора за отчет.период:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 3
        column = 2
        sheet.cell(row, column).value = f"Примечание, ВСЕГО начислений:"
        sheet.cell(row, column).style = style['style_8']

        row = row + 1
        column = 2
        sheet.cell(row, column).value = "Категория"
        sheet.cell(row, column).style = style['style_1_1']

        column = 3
        sheet.cell(row, column).value = "Сумма"
        sheet.cell(row, column).style = style['style_1_1']

        row = row + 1
        for item_exp in data['data_total_accrual_category']:

            column = 2
            sheet.cell(row, column).value = item_exp['category']
            sheet.cell(row, column).style = style['style_7']

            column = 3
            sheet.cell(row, column).value = item_exp['category_total_accrual_summa']
            sheet.cell(row, column).style = style['style_7']

            row += 1

        row = row + 2
        column = 2
        sheet.cell(row, column).value = f"Дивиденды начисленные, ВСЕГО:"
        sheet.cell(row, column).style = style['style_8']


    except Exception as ex:
        return {
            "status": "error",
            "data": None,
            "details": f'Ошибка формирования отчета Excel. Ошибка на строке {ex}'
        }

    file = f'{path_main}/src/media/statistic/result/Статистика_по_инвестициям_на {data["date_2"]}_{current_date.strftime("%d%m%Y_%H%M%S")}.xlsx'
    book_template.save(file)

    return {
        'status': 'success',
        'data': None,
        'details': 'Отчет Excel успешно сформирован'
    }