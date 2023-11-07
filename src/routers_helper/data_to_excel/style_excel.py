from openpyxl.styles import Font, NamedStyle, Side, Border, PatternFill, Alignment


def style_excel():
    # Стиль данных и ячеек
    style_main = NamedStyle(name="style_main")
    style_main.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_main.border = Border(left=side, right=side, top=side, bottom=side)
    style_main.fill = PatternFill("solid", fgColor="FFFFFF")
    style_main.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_1 = NamedStyle(name="style_1")
    style_1.font = Font(bold=True, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_1.border = Border(left=side, right=side, top=side, bottom=side)
    style_1.fill = PatternFill("solid", fgColor="FFFFFF")
    style_1.alignment = Alignment(horizontal='center', vertical='center')

    style_2 = NamedStyle(name="style_2")
    style_2.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_2.border = Border(left=side, right=side, top=side, bottom=side)
    style_2.fill = PatternFill("solid", fgColor="FFFFFF")
    style_2.alignment = Alignment(horizontal='center', vertical='center')

    style_3 = NamedStyle(name="style_3")
    style_3.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_3.border = Border(left=side, right=side, top=side, bottom=side)
    style_3.fill = PatternFill("solid", fgColor="b4c7e7")
    style_3.alignment = Alignment(horizontal='center', vertical='center')

    style_4 = NamedStyle(name="style_4")
    style_4.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_4.border = Border(left=side, right=side, top=side, bottom=side)
    style_4.fill = PatternFill("solid", fgColor="afd095")
    style_4.alignment = Alignment(horizontal='center', vertical='center')

    style_5 = NamedStyle(name="style_5")
    style_5.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_5.border = Border(left=side, right=side, top=side, bottom=side)
    style_5.fill = PatternFill("solid", fgColor="ffe699")
    style_5.alignment = Alignment(horizontal='center', vertical='center')

    style_6 = NamedStyle(name="style_6")
    style_6.font = Font(bold=False, size=12, color="000000")
    side = Side(style='thin', color="000000")
    style_6.border = Border(left=side, right=side, top=side, bottom=side)
    style_6.fill = PatternFill("solid", fgColor="f8cbad")
    style_6.alignment = Alignment(horizontal='center', vertical='center')

    style_7 = NamedStyle(name="style_7")
    style_7.font = Font(bold=False, size=10, color="000000")
    side = Side(style='thin', color="000000")
    style_7.border = Border(left=side, right=side, top=side, bottom=side)
    style_7.fill = PatternFill("solid", fgColor="FFFFFF")
    style_7.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_8 = NamedStyle(name="style_8")
    style_8.font = Font(bold=True, size=12, color="000000")
    style_8.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    style_9 = NamedStyle(name="style_9")
    style_9.font = Font(bold=False, size=12, color="000000")
    style_9.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    return {
        'style_main': style_main,
        'style_1': style_1,
        'style_2': style_2,
        'style_3': style_3,
        'style_4': style_4,
        'style_5': style_5,
        'style_6': style_6,
        'style_7': style_7,
        'style_8': style_8,
        'style_9': style_9,
    }