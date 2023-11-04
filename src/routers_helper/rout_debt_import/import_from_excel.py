from transliterate import translit
import re
import openpyxl
import pandas as pd
from fastapi import APIRouter, UploadFile

from src.config import path_main


# Импортировать заголовки столбцов реестра в Excel
router_import_headers_excel = APIRouter(
    prefix="/v1/ImportHeadersExcel",
    tags=["Import_to_database"]
)



@router_import_headers_excel.post("/")
async def import_headers(registry_debt_file: UploadFile):

    print(registry_debt_file.filename)

    with open(f'{path_main}/src/media/data/format_file.xlsx', 'wb+') as f:
        for chunk in registry_debt_file.file:
            f.write(chunk)

    path_file = f'{path_main}/src/media/data/format_file.xlsx'

    excel_data = pd.read_excel(path_file)
    heading = list(excel_data.columns.values)

    count = 0
    result = []
    for item in heading:
        count += 1
        result.append({
            "value": f"{count}",
            "text": item,
            "disabled": False
        })

    return result




'''
1. heading_transliterate() - Изменение заголовков столбцов, transliterat кирилицы в латиницу. 

ВАЖНО - текст в заголовке должен быть в одну строку, БЕЗ ПЕРЕНОСОВ
!!! Все ячейки с заголовками - ДОЛЖНЫ БЫТЬ ЗАПОЛНЕНЫ

# Полученные заголовки необходимо отформатировать:
# - запрещены символы ()/'%
# - заменить: / на _
# - заменить: % на percent
'''

def heading_transliterate():
    wookbook = openpyxl.load_workbook('/media/maks/Новый том/Python/work/fast_api/pravilo_crm/src/media/data/Реестр Рубль_ТЕСТ.xlsx')
    worksheet = wookbook.active
    for i in range(1):
        for col in worksheet.iter_cols(1, 60):
            ru_text = col[i].value
            text = translit(ru_text, language_code='ru', reversed=True)
            text_export = re.sub(' ', '_', text)

            print(text_export)

# heading_transliterate()


#Заготовка функции для извлечения данных из Excel
def export_ep():
    # with open(f'./data/payments_file.xlsx', 'wb+') as f:
    #     for chunk in file_object['payments_file'].chunks():
    #         f.write(chunk)

    path_file = f'{path_main}/src/media/data/Реестр ДИ по ЧС 4 рабочий для crm_1.xlsx'

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    for row in range(1, sheet.max_row):
        for cell in range(1, sheet.max_column):
            print(sheet[row][cell].value)



# export_ep()