import requests
from urllib.parse import quote
import json
import fake_useragent
import openpyxl
import re
from src.routers_helper.data_to_excel.bancrupt_to_excel import bankrupt_to_excel

HEADERS = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36', 'accept': '*/*'}      #User_Agent генерируется, чтобы сайт понимал, что к ниму обращается человек, а не бот
HOST = 'https://bankrot.fedresurs.ru/'
user = fake_useragent.UserAgent().random            #Можно рандомной генерации User_Agent

header = {
    'user-agent': user
}

def parse():

    # test_bankrupt.xlsx
    # Reestr_na_proverku.xlsx
    path_file = '/media/maks/Новый том/Python/work/fast_api/pravilo_crm/src/media/data/Reestr_na_proverku.xlsx'

    data_xlsx = extract_data_xlsx(path_file)

    url_main = f'https://bankrot.fedresurs.ru/bankrupts'

    html_main = get_html(url_main)
    cookies = html_main.cookies

    result = []
    for item in data_xlsx:

        url = f'https://bankrot.fedresurs.ru/backend/prsnbankrupts?searchString={quote(item)}&limit=15&offset=0'
        my_referer = f'https://bankrot.fedresurs.ru/bankrupts?searchString={quote(item)}'

        r = requests.get(url, cookies=cookies, headers={'referer': my_referer})
        dict_data = json.loads(r.text)

        if len(dict_data['pageData']) > 0:
            for data in dict_data['pageData']:
                # print(f'{data=}')

                guid = f'https://fedresurs.ru/backend/persons/{data["guid"]}'
                guid_referer = f'https://fedresurs.ru/person/{data["guid"]}'
                gdx = requests.get(guid, headers={'referer': guid_referer})
                guid_data = json.loads(gdx.text)


                try:
                    name_histories = guid_data['info']['nameHistories'][0]
                except:
                    name_histories = ''

                try:
                    inn = guid_data['info']['inn']
                except:
                    inn = ''

                try:
                    snils = guid_data['info']['snils']
                except:
                    snils = ''

                try:
                    address = guid_data['info']['address']
                except:
                    address = ''

                try:
                    number_legal_cases = guid_data['legalCases'][0]['number']
                except:
                    number_legal_cases = ''

                try:
                    tribunal = guid_data['legalCases'][0]['courtName']
                except:
                    tribunal = ''

                result.append({
                    'full_name': guid_data['info']['fullName'],
                    'name_histories': name_histories,
                    'birthdate_bankrupt': guid_data['info']['birthdateBankruptcy'],
                    'birth_place_bankrupt': guid_data['info']['birthplaceBankruptcy'],
                    'address': address,
                    'inn': inn,
                    'snils': snils,
                    'number_legal_cases': number_legal_cases,
                    'tribunal': tribunal})

    bankrupt_to_excel(result)

'''
Функция проверяет отвечает ли сайт на наш запрос.
Применяется если нет необходимости проходить авторизацию на сайте
'''
def get_html(url, params=None):
    d = requests.get(url, headers=HEADERS, params=params)
    return d


def extract_data_xlsx(path_file):
    result = []

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    num_cel = 0
    for row in range(1, (sheet.max_row + 1)):
        for cell in range(1, sheet.max_column):
            if re.findall(r'inn' or r'fio', str(sheet[row][cell].value)):
                num_cel = cell
        data_cel = sheet[row][num_cel].value
        try:
            data_cel = re.sub(r'^[\s-]+|\s+$|\n+|^\s+', '', data_cel)
        except:
            data_cel = data_cel
        result.append(str(data_cel))

    return result

# parse()