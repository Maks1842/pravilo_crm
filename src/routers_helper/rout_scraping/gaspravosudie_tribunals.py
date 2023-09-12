import requests
from subprocess import check_output
from urllib.parse import quote
import urllib.request
import json
from bs4 import BeautifulSoup
import fake_useragent
import openpyxl
import re

from fastapi import APIRouter, Depends
from sqlalchemy import select, insert, func, distinct, update, desc, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from src.database import get_async_session

from src.references.models import ref_tribunal

HEADERS = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36', 'accept': '*/*'}      #User_Agent генерируется, чтобы сайт понимал, что к ниму обращается человек, а не бот
HOST = 'https://bankrot.fedresurs.ru/'
user = fake_useragent.UserAgent().random            #Можно рандомной генерации User_Agent

header = {
    'user-agent': user
}


async def parse_gas_ms(session):

    path_file = '/home/maks/Загрузки/index.php'

    with open(path_file) as file:
        text = str(file.read())

    class_code_list = re.findall(r"(?<=\[balloons_user\[').+(?='\]\.)", text)
    tribunal_list = re.findall(r"(?<=name:').+(?=',adress)", text)
    address_tribun_list = re.findall(r"(?<=adress:').+(?=',coord)", text)

    count = 0

    for item in class_code_list:

        try:
            tribunal = {
                'class_code': item,
                'name': tribunal_list[count],
                'address': address_tribun_list[count]}
            count += 1

            post_data = insert(ref_tribunal).values(tribunal)

            await session.execute(post_data)
            await session.commit()

            print('успех')

            return
        except Exception as ex:
            print(f'{item} {ex}')
            return
    print(count)


    return


def parse():

    data_region = [26,]

    url_main = f'https://ej.sudrf.ru/api/appeal/getCourtsInRegion'

    html_main = get_html(url_main)
    cookies = html_main.cookies

    result = []
    for item in data_region:

        url = f'https://ej.sudrf.ru/api/appeal/getCourtsInRegion?regionCode={item}'
        my_referer = f'https://ej.sudrf.ru/appeal?type=217.01&process_id=e0a0b464-f144-4844-9e64-d7612db04e8c'
        print(url)
        print(my_referer)


        r = requests.get(url, cookies=cookies, headers={'referer': my_referer})
        print(r)
        dict_data = r.content
        print(dict_data)


    #     dict_data = json.loads(r.text)
    #
    #     if len(dict_data['pageData']) > 0:
    #         for data in dict_data['pageData']:
    #             # print(f'{data=}')
    #
    #             guid = f'https://fedresurs.ru/backend/persons/{data["guid"]}'
    #             guid_referer = f'https://fedresurs.ru/person/{data["guid"]}'
    #             gdx = requests.get(guid, headers={'referer': guid_referer})
    #             guid_data = json.loads(gdx.text)
    #
    #
    #             try:
    #                 name_histories = guid_data['info']['nameHistories'][0]
    #             except:
    #                 name_histories = ''
    #
    #             try:
    #                 inn = guid_data['info']['inn']
    #             except:
    #                 inn = ''
    #
    #             try:
    #                 snils = guid_data['info']['snils']
    #             except:
    #                 snils = ''
    #
    #             try:
    #                 address = guid_data['info']['address']
    #             except:
    #                 address = ''
    #
    #             try:
    #                 number_legal_cases = guid_data['legalCases'][0]['number']
    #             except:
    #                 number_legal_cases = ''
    #
    #             try:
    #                 tribunal = guid_data['legalCases'][0]['courtName']
    #             except:
    #                 tribunal = ''
    #
    #             result.append({
    #                 'full_name': guid_data['info']['fullName'],
    #                 'name_histories': name_histories,
    #                 'birthdate_bankrupt': guid_data['info']['birthdateBankruptcy'],
    #                 'birth_place_bankrupt': guid_data['info']['birthplaceBankruptcy'],
    #                 'address': address,
    #                 'inn': inn,
    #                 'snils': snils,
    #                 'number_legal_cases': number_legal_cases,
    #                 'tribunal': tribunal})
    #
    # bankrupt_to_excel(result)

'''
Функция проверяет отвечает ли сайт на наш запрос.
Применяется если нет необходимости проходить авторизацию на сайте
'''
def get_html(url, params=None):
    d = requests.get(url, headers=HEADERS, params=params)
    return d


def extract_data_xlsx(path_file):
    result = []

    book = openpyxl.load_workbook(path_file)
    sheet = book.active

    num_cel = 0
    for row in range(1, (sheet.max_row + 1)):
        for cell in range(1, sheet.max_column):
            if re.findall(r'inn' or r'fio', str(sheet[row][cell].value)):
                num_cel = cell
        data_cel = sheet[row][num_cel].value
        try:
            data_cel = re.sub(r'^[\s-]+|\s+$|\n+|^\s+', '', data_cel)
        except:
            data_cel = data_cel
        result.append(str(data_cel))

    return result

# parse_gas_ms()
# parse()